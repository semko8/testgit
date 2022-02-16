import requests
from openpyxl import load_workbook
from urllib.request import urlopen
import json


filename1 = r"C:\Users\user\my_ip_list.xlsx"
workbook2 = load_workbook(filename=filename1)
sheet1 = workbook2["Domains"]   # sheet name

# security trails api endpoint
url = "https://api.securitytrails.com/v1/domains/list" 
querystring = {"include_ips":"true","scroll":"true"}

# security trails free api limit 50
for i in range(1,50):
    domain = ""
    
    # get ip 
    get_cell = sheet1.cell(row=i+1, column=1).value

    payload = {"filter": {"ipv4": get_cell}}
    
    # enter api key 
    headers = {
        "Content-Type": "application/json",
        "APIKEY": "<my-api-key>"
    }
    
    # get response
    response = requests.request("POST", url, json=payload, headers=headers, params=querystring)
    
    try:
        text_response = json.loads(response.text)
        
        for a in range(len(text_response["records"])):
            domain += text_response["records"][a]["hostname"] + "\n"
        # write records (domain names) to xlsx file
        sheet1.cell(row=i+1, column=2).value = domain
        workbook2.save(filename=filename1)    
    except:
        print(get_cell + " returns no response data")

    print(response.text)